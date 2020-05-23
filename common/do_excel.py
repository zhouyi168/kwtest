#!usr/bin/env python
#-*- coding:utf-8 _*-
"""
@author:末夏
@file: do_excel.py
@time: 2020/04/14
"""
from openpyxl import load_workbook

class Case:
    id = None
    title = None
    action = None
    locate_type = None
    expression = None
    value = None
    excu_time = None
    result = None
    error_info = None
    error_img = None

class DoExcel:
    def __init__(self,filename):
        self.filename=filename

    def readrow(self,sheetname,id):
        case=Case()
        workbook = load_workbook(self.filename)
        table = workbook[sheetname]
        maxrow=table.max_row
        for row in range(1,maxrow+1):
            if row==id+1:
                case.id = table.cell(row=row, column=1).value
                case.title = table.cell(row=row, column=2).value
                case.action = table.cell(row=row, column=3).value
                case.locate_type = table.cell(row=row, column=4).value
                case.expression = table.cell(row=row, column=5).value
                case.value = table.cell(row=row, column=6).value
                case.excu_time = table.cell(row=row, column=7).value
                case.result = table.cell(row=row, column=8).value
                case.error_info = table.cell(row=row, column=9).value
                case.error_img = table.cell(row=row, column=10).value
                break
        return case

    def readexcel(self,sheetname):
        data=[]
        workbook=load_workbook(self.filename)
        table=workbook[sheetname]
        maxrow=table.max_row
        for row in range(2,maxrow+1):
            case = Case()
            case.id=table.cell(row=row,column=1).value
            case.title=table.cell(row=row,column=2).value
            case.action = table.cell(row=row, column=3).value
            case.locate_type = table.cell(row=row, column=4).value
            case.expression = table.cell(row=row, column=5).value
            case.value=table.cell(row=row, column=6).value
            case.excu_time = table.cell(row=row, column=7).value
            case.result = table.cell(row=row, column=8).value
            case.error_info = table.cell(row=row, column=9).value
            case.error_img = table.cell(row=row, column=10).value
            data.append(case)
        workbook.close()
        return data

    def write_excel(self,sheetname,row,col,value):
        workbook = load_workbook(self.filename)
        table = workbook[sheetname]
        maxrow=table.max_row
        maxcol=table.max_column
        try:
            for r in range(2,maxrow+1):
                if row == r:
                    for c in range(1,maxcol+1):
                         if c==col:
                             table.cell(row,col).value=value
                             break
                    break
            workbook.save(self.filename)
            workbook.close()
        except Exception as e:
            print('写入{}表出错'.format(sheetname))
            raise e

    def maxrows(self,sheetname):
        workbook = load_workbook(self.filename)
        table = workbook[sheetname]
        maxrow = table.max_row
        return maxrow

if __name__=="__main__":
    import os
    from common import con_path
    file_path=con_path.testdata_path
    # data=DoExcel(file_path).readexcel('baidu')
    # for case in data:
    #     print(case.title)
    # DoExcel(file_path).write_excel('baidu',3,8,'pass')
    data=DoExcel(file_path).readrow('baidu',2)
    print(data.title)






